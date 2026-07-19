#!/usr/bin/env python3
"""Contrato e validacoes compartilhadas da importacao de cursos."""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


SCHEMA_VERSION = 1

SHEET_COLUMNS = {
    "metadados": ("chave", "valor"),
    "cursos": ("sigla", "curso", "regime"),
    "docentes": ("docente", "apelido", "unidade", "subunidade"),
    "componentes": (
        "sigla",
        "periodo",
        "codigo",
        "cor",
        "componente",
        "abreviacao",
        "ch",
    ),
    "turmas": ("sigla", "ano", "turno"),
    "salas": ("campus", "codigo", "nome", "capacidade", "tipo", "recursos"),
}

REQUIRED_SHEETS = ("metadados", "cursos", "componentes", "turmas")
IMPORT_SHEETS = ("cursos", "docentes", "componentes", "turmas", "salas")
HEX_COLOR_RE = re.compile(r"^#[0-9A-Fa-f]{6}$")
COURSE_CODE_RE = re.compile(r"^[A-Z0-9]{2,10}$")


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFD", normalize_text(value).casefold())
    return "".join(char for char in text if unicodedata.category(char) != "Mn")


def normalize_header(value: Any) -> str:
    return normalize_key(value).replace(" ", "_").replace("-", "_")


def parse_integer(value: Any) -> int | None:
    if value is None or normalize_text(value) == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else None


def row_is_empty(values: Iterable[Any]) -> bool:
    return all(normalize_text(value) == "" for value in values)


@dataclass(frozen=True)
class ImportIssue:
    severity: str
    code: str
    message: str
    sheet: str = ""
    row: int | None = None
    field: str = ""


@dataclass
class ImportValidation:
    source_file: str
    course_sigla: str
    rows: dict[str, list[dict[str, Any]]]
    issues: list[ImportIssue]

    @property
    def errors(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "error"]

    @property
    def warnings(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "warning"]

    @property
    def valid(self) -> bool:
        return not self.errors

    def to_report(self) -> dict[str, Any]:
        return {
            "schema_version": SCHEMA_VERSION,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_file": self.source_file,
            "course_sigla": self.course_sigla,
            "valid": self.valid,
            "counts": {sheet: len(self.rows.get(sheet, [])) for sheet in IMPORT_SHEETS},
            "summary": {
                "errors": len(self.errors),
                "warnings": len(self.warnings),
            },
            "issues": [asdict(issue) for issue in self.issues],
        }


def read_sheet_rows(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    issues: list[ImportIssue],
) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        if sheet_name in REQUIRED_SHEETS:
            issues.append(
                ImportIssue("error", "missing_sheet", f"Aba obrigatoria ausente: {sheet_name}.", sheet_name)
            )
        return []

    worksheet = workbook[sheet_name]
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [normalize_header(value) for value in header_values]
    expected = SHEET_COLUMNS[sheet_name]
    missing = [column for column in expected if column not in headers]
    if missing:
        issues.append(
            ImportIssue(
                "error",
                "missing_columns",
                f"Colunas obrigatorias ausentes: {', '.join(missing)}.",
                sheet_name,
                1,
            )
        )
        return []

    index_by_column = {column: headers.index(column) for column in expected}
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_is_empty(values):
            continue
        item = {column: values[index] if index < len(values) else None for column, index in index_by_column.items()}
        item["_row"] = row_number
        rows.append(item)
    return rows


def _add_required_text_issues(
    rows: list[dict[str, Any]],
    sheet: str,
    fields: tuple[str, ...],
    issues: list[ImportIssue],
) -> None:
    for row in rows:
        for field in fields:
            if not normalize_text(row.get(field)):
                issues.append(
                    ImportIssue(
                        "error",
                        "required_value",
                        f"O campo '{field}' e obrigatorio.",
                        sheet,
                        row["_row"],
                        field,
                    )
                )


def _add_duplicate_issues(
    rows: list[dict[str, Any]],
    sheet: str,
    key_fields: tuple[str, ...],
    issues: list[ImportIssue],
) -> None:
    first_row_by_key: dict[tuple[str, ...], int] = {}
    for row in rows:
        key = tuple(normalize_key(row.get(field)) for field in key_fields)
        if not all(key):
            continue
        if key in first_row_by_key:
            issues.append(
                ImportIssue(
                    "error",
                    "duplicate_row",
                    f"Registro duplicado; primeira ocorrencia na linha {first_row_by_key[key]}.",
                    sheet,
                    row["_row"],
                    ",".join(key_fields),
                )
            )
        else:
            first_row_by_key[key] = row["_row"]


def _read_base_context(base_path: Path | None) -> dict[str, Any]:
    context = {
        "course_codes": set(),
        "teacher_names": set(),
        "component_keys": set(),
        "cohort_keys": set(),
        "room_keys": set(),
        "regimes": set(),
        "shifts": set(),
    }
    if base_path is None or not base_path.exists():
        return context

    workbook = openpyxl.load_workbook(base_path, data_only=True, read_only=True)
    sheet_fields = {
        "cursos": ("sigla",),
        "docentes": ("docente",),
        "componentes": ("sigla", "codigo", "periodo"),
        "turmas": ("sigla", "ano"),
        "salas": ("campus", "codigo"),
        "horarios": ("regime", "turno"),
    }
    for sheet_name, fields in sheet_fields.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = [normalize_header(cell.value) for cell in worksheet[1]]
        indexes = {field: headers.index(field) for field in fields if field in headers}
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if row_is_empty(values):
                continue
            key = tuple(normalize_key(values[indexes[field]]) for field in fields if field in indexes)
            if not all(key):
                continue
            if sheet_name == "cursos":
                context["course_codes"].add(key[0])
            elif sheet_name == "docentes":
                context["teacher_names"].add(key[0])
            elif sheet_name == "componentes":
                context["component_keys"].add(key)
            elif sheet_name == "turmas":
                context["cohort_keys"].add(key)
            elif sheet_name == "salas":
                context["room_keys"].add(key)
            elif sheet_name == "horarios":
                context["regimes"].add(key[0])
                context["shifts"].add(key[1])
    return context


def validate_course_import(source_path: Path, base_path: Path | None = None) -> ImportValidation:
    issues: list[ImportIssue] = []
    try:
        workbook = openpyxl.load_workbook(source_path, data_only=True)
    except Exception as exc:
        return ImportValidation(
            source_path.name,
            "",
            {sheet: [] for sheet in IMPORT_SHEETS},
            [ImportIssue("error", "invalid_workbook", f"Nao foi possivel abrir a planilha: {exc}")],
        )

    rows = {
        sheet: read_sheet_rows(workbook, sheet, issues)
        for sheet in SHEET_COLUMNS
    }

    metadata = {
        normalize_key(row.get("chave")): normalize_text(row.get("valor"))
        for row in rows["metadados"]
        if normalize_text(row.get("chave"))
    }
    version = parse_integer(metadata.get("schema_version"))
    if version != SCHEMA_VERSION:
        issues.append(
            ImportIssue(
                "error",
                "schema_version",
                f"schema_version deve ser {SCHEMA_VERSION}; encontrado: {metadata.get('schema_version', 'vazio')}.",
                "metadados",
            )
        )

    courses = rows["cursos"]
    if len(courses) != 1:
        issues.append(
            ImportIssue(
                "error",
                "course_count",
                f"A planilha deve cadastrar exatamente um curso; encontrados: {len(courses)}.",
                "cursos",
            )
        )

    _add_required_text_issues(courses, "cursos", ("sigla", "curso", "regime"), issues)
    _add_required_text_issues(
        rows["docentes"], "docentes", ("docente", "unidade", "subunidade"), issues
    )
    _add_required_text_issues(
        rows["componentes"],
        "componentes",
        ("sigla", "periodo", "codigo", "cor", "componente", "abreviacao", "ch"),
        issues,
    )
    _add_required_text_issues(rows["turmas"], "turmas", ("sigla", "ano", "turno"), issues)
    _add_required_text_issues(
        rows["salas"], "salas", ("campus", "codigo", "nome", "capacidade", "tipo"), issues
    )

    if not rows["componentes"]:
        issues.append(ImportIssue("error", "empty_sheet", "Inclua ao menos uma componente.", "componentes"))
    if not rows["turmas"]:
        issues.append(ImportIssue("error", "empty_sheet", "Inclua ao menos uma turma.", "turmas"))

    course_sigla = normalize_text(courses[0].get("sigla")).upper() if len(courses) == 1 else ""
    if course_sigla and not COURSE_CODE_RE.fullmatch(course_sigla):
        issues.append(
            ImportIssue(
                "error",
                "invalid_course_code",
                "A sigla deve ter de 2 a 10 letras maiusculas ou numeros.",
                "cursos",
                courses[0]["_row"],
                "sigla",
            )
        )

    regime = parse_integer(courses[0].get("regime")) if len(courses) == 1 else None
    if courses and (regime is None or regime <= 0):
        issues.append(
            ImportIssue("error", "invalid_integer", "Regime deve ser um inteiro positivo.", "cursos", courses[0]["_row"], "regime")
        )

    for sheet in ("componentes", "turmas"):
        for row in rows[sheet]:
            sigla = normalize_text(row.get("sigla")).upper()
            if course_sigla and sigla != course_sigla:
                issues.append(
                    ImportIssue(
                        "error",
                        "course_mismatch",
                        f"A sigla deve ser '{course_sigla}', o curso deste arquivo.",
                        sheet,
                        row["_row"],
                        "sigla",
                    )
                )

    for row in rows["docentes"]:
        if len(normalize_text(row.get("docente")).split()) < 2:
            issues.append(
                ImportIssue(
                    "error",
                    "teacher_full_name",
                    "Informe o nome completo do docente.",
                    "docentes",
                    row["_row"],
                    "docente",
                )
            )

    for row in rows["componentes"]:
        workload = parse_integer(row.get("ch"))
        if workload is None or workload <= 0:
            issues.append(
                ImportIssue("error", "invalid_integer", "CH deve ser um inteiro positivo.", "componentes", row["_row"], "ch")
            )
        color = normalize_text(row.get("cor"))
        if color and not HEX_COLOR_RE.fullmatch(color):
            issues.append(
                ImportIssue("error", "invalid_color", "Cor deve usar o formato #RRGGBB.", "componentes", row["_row"], "cor")
            )

    for row in rows["turmas"]:
        year = parse_integer(row.get("ano"))
        if year is None or not 2000 <= year <= 2100:
            issues.append(
                ImportIssue("error", "invalid_year", "Ano deve estar entre 2000 e 2100.", "turmas", row["_row"], "ano")
            )

    for row in rows["salas"]:
        capacity = parse_integer(row.get("capacidade"))
        if capacity is None or capacity < 0:
            issues.append(
                ImportIssue("error", "invalid_capacity", "Capacidade deve ser um inteiro nao negativo.", "salas", row["_row"], "capacidade")
            )

    _add_duplicate_issues(rows["docentes"], "docentes", ("docente",), issues)
    _add_duplicate_issues(rows["componentes"], "componentes", ("sigla", "codigo", "periodo"), issues)
    _add_duplicate_issues(rows["turmas"], "turmas", ("sigla", "ano"), issues)
    _add_duplicate_issues(rows["salas"], "salas", ("campus", "codigo"), issues)

    base = _read_base_context(base_path)
    if course_sigla and normalize_key(course_sigla) in base["course_codes"]:
        issues.append(
            ImportIssue(
                "error",
                "course_exists",
                f"O curso '{course_sigla}' ja existe na planilha-base.",
                "cursos",
                courses[0]["_row"] if courses else None,
                "sigla",
            )
        )
    if regime is not None and base["regimes"] and normalize_key(regime) not in base["regimes"]:
        issues.append(
            ImportIssue(
                "error",
                "unknown_regime",
                f"O regime {regime} nao possui horarios na planilha-base.",
                "cursos",
                courses[0]["_row"] if courses else None,
                "regime",
            )
        )

    for row in rows["docentes"]:
        if normalize_key(row.get("docente")) in base["teacher_names"]:
            issues.append(
                ImportIssue(
                    "warning",
                    "teacher_exists",
                    "Docente ja existe na base e sera ignorado na incorporacao.",
                    "docentes",
                    row["_row"],
                    "docente",
                )
            )

    for row in rows["turmas"]:
        shift = normalize_key(row.get("turno"))
        if shift and base["shifts"] and shift not in base["shifts"]:
            issues.append(
                ImportIssue(
                    "warning",
                    "unknown_shift",
                    "Turno ainda nao aparece na grade de horarios; confirme se e uma combinacao intencional.",
                    "turmas",
                    row["_row"],
                    "turno",
                )
            )

    return ImportValidation(source_path.name, course_sigla, rows, issues)


def write_validation_report(validation: ImportValidation, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(validation.to_report(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

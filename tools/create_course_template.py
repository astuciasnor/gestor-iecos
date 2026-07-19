#!/usr/bin/env python3
"""Gera a planilha modelo para cadastro estruturado de um novo curso."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from course_import_common import SCHEMA_VERSION, SHEET_COLUMNS


HEADER_FILL = PatternFill("solid", fgColor="006A6A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="E7F3F3")
THIN_BORDER = Border(bottom=Side(style="thin", color="B8C6C6"))

SHEET_WIDTHS = {
    "metadados": (24, 36),
    "cursos": (14, 42, 12),
    "docentes": (42, 20, 18, 22),
    "componentes": (12, 12, 18, 14, 44, 20, 12),
    "turmas": (12, 12, 20),
    "salas": (20, 16, 32, 16, 24, 44),
}


def _style_table_sheet(worksheet, columns: tuple[str, ...]) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(columns)).coordinate}"
    worksheet.row_dimensions[1].height = 26
    for index, column in enumerate(columns, start=1):
        cell = worksheet.cell(1, index, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        worksheet.column_dimensions[cell.column_letter].width = SHEET_WIDTHS[worksheet.title][index - 1]
    for row_number in range(2, 202):
        for column_number in range(1, len(columns) + 1):
            worksheet.cell(row_number, column_number).border = THIN_BORDER


def _add_validations(workbook: Workbook) -> None:
    regime = DataValidation(type="list", formula1='"1,2"', allow_blank=False)
    workbook["cursos"].add_data_validation(regime)
    regime.add("C2:C20")

    workload = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=False)
    workbook["componentes"].add_data_validation(workload)
    workload.add("G2:G500")

    year = DataValidation(type="whole", operator="between", formula1="2000", formula2="2100", allow_blank=False)
    workbook["turmas"].add_data_validation(year)
    year.add("B2:B200")

    capacity = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
    workbook["salas"].add_data_validation(capacity)
    capacity.add("D2:D500")

    invalid_color_fill = PatternFill("solid", fgColor="FFC7CE")
    workbook["componentes"].conditional_formatting.add(
        "D2:D500",
        FormulaRule(
            formula=['AND(D2<>"",OR(LEFT(D2,1)<>"#",LEN(D2)<>7))'],
            fill=invalid_color_fill,
        ),
    )


def build_template() -> Workbook:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "instrucoes"
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100
    instructions["A1"] = "Cadastro de novo curso"
    instructions["A1"].font = Font(size=18, bold=True, color="006A6A")
    instructions.merge_cells("A1:B1")

    guidance = [
        ("Como usar", "Preencha somente as abas de dados. Nao altere os nomes das abas nem dos cabecalhos."),
        ("Um curso por arquivo", "A aba cursos deve conter exatamente uma linha."),
        ("Sigla", "Repita a mesma sigla do curso nas abas componentes e turmas."),
        ("Componentes", "A chave e sigla + codigo + periodo. O mesmo codigo pode existir em periodos diferentes."),
        ("Docentes", "Inclua apenas docentes necessarios ao novo curso. Docentes ja existentes serao reconhecidos."),
        ("Salas", "Separe recursos por ponto e virgula, por exemplo: Projetor; Ar-condicionado."),
        ("Cores", "Use hexadecimal no formato #RRGGBB."),
        ("Validar", "Execute: python tools/validate_course_import.py CAMINHO_DA_PLANILHA"),
        ("Previa", "Execute: python tools/merge_course_import.py CAMINHO_DA_PLANILHA"),
        ("Incorporar", "Somente apos revisar a previa, use --apply --confirm SIGLA."),
    ]
    for row_number, (title, text) in enumerate(guidance, start=3):
        instructions.cell(row_number, 1, title).font = Font(bold=True, color="29485C")
        instructions.cell(row_number, 1).fill = SECTION_FILL
        instructions.cell(row_number, 2, text).alignment = Alignment(wrap_text=True, vertical="top")
        instructions.row_dimensions[row_number].height = 34

    examples = [
        ("Exemplo cursos", "ZZ | Engenharia de Exemplo | 1"),
        ("Exemplo componente", "ZZ | I | ZZ01001 | #9BF6FF | Introducao | INTRO | 60"),
        ("Exemplo turma", "ZZ | 2026 | Manha"),
        ("Exemplo sala", "BRAGANCA | S01 | Sala 01 | 40 | Sala de aula | Projetor"),
    ]
    start_row = len(guidance) + 5
    for offset, (title, text) in enumerate(examples):
        row_number = start_row + offset
        instructions.cell(row_number, 1, title).font = Font(bold=True, color="29485C")
        instructions.cell(row_number, 2, text).font = Font(name="Consolas", size=10)

    for sheet_name, columns in SHEET_COLUMNS.items():
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(columns)
        _style_table_sheet(worksheet, columns)

    workbook["metadados"].append(("schema_version", SCHEMA_VERSION))
    workbook["metadados"].append(("responsavel_nome", ""))
    workbook["metadados"].append(("responsavel_email", ""))
    workbook["metadados"].append(("observacoes", ""))
    _add_validations(workbook)
    return workbook


def main() -> int:
    root_dir = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description="Gera a planilha modelo de cadastro de curso.")
    parser.add_argument(
        "--output",
        default=str(root_dir / "dados" / "importacoes" / "modelo_cadastro_curso.xlsx"),
        help="Caminho do arquivo .xlsx a gerar",
    )
    parser.add_argument("--overwrite", action="store_true", help="Sobrescreve um modelo existente")
    args = parser.parse_args()

    output_path = Path(args.output).expanduser().resolve()
    if output_path.exists() and not args.overwrite:
        parser.error(f"arquivo ja existe: {output_path}; use --overwrite para substitui-lo")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    build_template().save(output_path)
    print(f"Modelo criado: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
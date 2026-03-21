#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
convert_data.py

Planilha esperada:

docentes:
  docente | unidade | subunidade

componentes:
  sigla | periodo | codigo | cor | componente | abreviacao | ch

turmas:
  sigla | ano | turno

cursos:
  sigla | curso

horarios:
  turno | ordem | faixa

feriados:
  data | feriado | dia | tipo

periodos_letivos:
  ano | periodo_letivo | inicio | fim

Requisitos:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils.datetime import from_excel


def norm_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace(" ", "_").replace("-", "_")
    return "".join(ch for ch in text if ch.isalnum() or ch == "_")


def to_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return 0


def to_date_iso(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt.date().isoformat()
            if isinstance(dt, date):
                return dt.isoformat()
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return ""

    if len(text) >= 10 and text[2] == "/" and text[5] == "/":
        dd, mm, yyyy = text[:2], text[3:5], text[6:10]
        return f"{yyyy}-{mm}-{dd}"

    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]

    return text


def is_intervalo(faixa: str) -> bool:
    upper = faixa.upper()
    if "INTERVALO" in upper:
        return True

    try:
        start_text, end_text = [part.strip() for part in faixa.split("-")]
        h1, m1 = [int(part) for part in start_text.split(":")]
        h2, m2 = [int(part) for part in end_text.split(":")]
        duration = (h2 * 60 + m2) - (h1 * 60 + m1)
        return 1 <= duration <= 25
    except Exception:
        return False


@dataclass(frozen=True)
class SheetSpec:
    name: str
    required_cols: List[str]
    optional: bool = False


def read_sheet(wb: openpyxl.Workbook, spec: SheetSpec) -> List[Dict[str, Any]]:
    if spec.name not in wb.sheetnames:
        if spec.optional:
            return []
        raise ValueError(f"Aba '{spec.name}' nao existe na planilha.")

    ws = wb[spec.name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [norm_key(header) for header in header_row if header is not None]

    missing = [col for col in spec.required_cols if col not in headers]
    if missing:
        raise ValueError(
            f"Aba '{spec.name}': faltando colunas: {missing}. Encontradas: {headers}"
        )

    idx_map: Dict[int, str] = {}
    for idx, header in enumerate([norm_key(header) for header in header_row]):
        if header:
            idx_map[idx] = header

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        item: Dict[str, Any] = {}
        for idx, value in enumerate(row):
            col = idx_map.get(idx)
            if not col:
                continue
            item[col] = value
        rows.append(item)

    return rows


def build_json_from_excel(xlsx_path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    specs = [
        SheetSpec("docentes", ["docente", "unidade", "subunidade"]),
        SheetSpec("componentes", ["sigla", "periodo", "codigo", "cor", "componente", "abreviacao", "ch"]),
        SheetSpec("turmas", ["sigla", "ano", "turno"]),
        SheetSpec("cursos", ["sigla", "curso"]),
        SheetSpec("horarios", ["turno", "ordem", "faixa"]),
        SheetSpec("feriados", ["data", "feriado", "dia", "tipo"]),
        SheetSpec("periodos_letivos", ["ano", "periodo_letivo", "inicio", "fim"], optional=True),
    ]

    raw: Dict[str, List[Dict[str, Any]]] = {}
    for spec in specs:
        raw[spec.name] = read_sheet(wb, spec)

    data: Dict[str, Any] = {
        "meta": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_file": xlsx_path.name,
        }
    }

    data["docentes"] = [
        {
            "docente": to_str(row.get("docente")),
            "unidade": to_str(row.get("unidade")),
            "subunidade": to_str(row.get("subunidade")),
        }
        for row in raw["docentes"]
    ]

    data["cursos"] = [
        {
            "sigla": to_str(row.get("sigla")),
            "curso": to_str(row.get("curso")),
        }
        for row in raw["cursos"]
    ]

    data["componentes"] = [
        {
            "sigla": to_str(row.get("sigla")),
            "periodo": to_str(row.get("periodo")),
            "codigo": to_str(row.get("codigo")),
            "cor": to_str(row.get("cor")),
            "componente": to_str(row.get("componente")),
            "abreviacao": to_str(row.get("abreviacao")),
            "ch": to_int(row.get("ch")),
        }
        for row in raw["componentes"]
    ]

    turmas_out = []
    for row in raw["turmas"]:
        sigla = to_str(row.get("sigla"))
        ano = to_int(row.get("ano"))
        turno = to_str(row.get("turno"))
        turma_id = f"{sigla}{ano}" if sigla and ano else ""
        turmas_out.append(
            {
                "sigla": sigla,
                "ano": ano,
                "turno": turno,
                "turma_id": turma_id,
                "turma_label": turma_id,
            }
        )
    data["turmas"] = turmas_out

    horarios_out = []
    for row in raw["horarios"]:
        turno = to_str(row.get("turno"))
        ordem = to_int(row.get("ordem"))
        faixa = to_str(row.get("faixa"))
        if not turno or not faixa:
            continue
        faixa_out = f"INTERVALO ({faixa})" if is_intervalo(faixa) else faixa
        horarios_out.append({"turno": turno, "ordem": ordem, "faixa": faixa_out})

    horarios_out.sort(key=lambda item: (item["turno"], item["ordem"]))
    data["horarios"] = horarios_out

    horarios_por_turno: Dict[str, List[str]] = {}
    for horario in horarios_out:
        horarios_por_turno.setdefault(horario["turno"], []).append(horario["faixa"])
    data["horarios_por_turno"] = horarios_por_turno

    data["feriados"] = [
        {
            "data": to_date_iso(row.get("data")),
            "feriado": to_str(row.get("feriado")),
            "dia": to_str(row.get("dia")),
            "tipo": to_str(row.get("tipo")),
        }
        for row in raw["feriados"]
    ]

    periodos_letivos = []
    for row in raw["periodos_letivos"]:
        ano = to_int(row.get("ano"))
        periodo_letivo = to_str(row.get("periodo_letivo")).upper()
        inicio = to_date_iso(row.get("inicio"))
        fim = to_date_iso(row.get("fim"))
        if not periodo_letivo or not inicio or not fim:
            continue
        periodos_letivos.append(
            {
                "ano": ano,
                "periodo_letivo": periodo_letivo,
                "inicio": inicio,
                "fim": fim,
                "label": f"{ano} - {periodo_letivo}" if ano else periodo_letivo,
            }
        )

    periodos_letivos.sort(key=lambda item: (item.get("inicio", ""), item.get("periodo_letivo", "")))
    data["periodos_letivos"] = periodos_letivos

    return data


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    root_dir = script_dir.parent

    default_input = root_dir / "dados" / "planilha_base.xlsx"
    default_output = root_dir / "dados_app.json"

    parser = argparse.ArgumentParser(description="Converte Excel (.xlsx) em JSON para o app.")
    parser.add_argument("--input", default=str(default_input), help="Caminho do .xlsx")
    parser.add_argument("--output", default=str(default_output), help="Caminho do .json de saida")
    args = parser.parse_args()

    in_path = Path(args.input).expanduser().resolve()
    out_path = Path(args.output).expanduser().resolve()

    if not in_path.exists():
        raise SystemExit(
            f"ERRO: Arquivo nao encontrado: {in_path}\n"
            "Coloque o arquivo na pasta 'dados/planilha_base.xlsx' ou especifique o caminho com --input"
        )

    data = build_json_from_excel(in_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK: gerado {out_path}")
    print("Chaves:", ", ".join(data.keys()))


if __name__ == "__main__":
    main()

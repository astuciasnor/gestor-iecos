#!/usr/bin/env python3
"""Incorpora um curso validado na planilha-base com previa e backup."""

from __future__ import annotations

import argparse
import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from convert_data import build_json_from_excel
from course_import_common import (
    IMPORT_SHEETS,
    SHEET_COLUMNS,
    ImportValidation,
    normalize_header,
    normalize_key,
    normalize_text,
    parse_integer,
    validate_course_import,
    write_validation_report,
)


def _clean_value(sheet: str, field: str, value: Any) -> Any:
    if field in {"regime", "ch", "ano", "capacidade"}:
        return parse_integer(value)
    text = normalize_text(value)
    if field == "sigla":
        return text.upper()
    if field == "cor":
        return text.upper()
    return text


def _ensure_sheet(workbook: openpyxl.Workbook, sheet: str):
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    worksheet = workbook.create_sheet(sheet)
    worksheet.append(SHEET_COLUMNS[sheet])
    return worksheet


def _existing_keys(worksheet, fields: tuple[str, ...]) -> set[tuple[str, ...]]:
    headers = [normalize_header(cell.value) for cell in worksheet[1]]
    indexes = {field: headers.index(field) for field in fields if field in headers}
    keys: set[tuple[str, ...]] = set()
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        key = tuple(
            normalize_key(values[indexes[field]])
            for field in fields
            if field in indexes and indexes[field] < len(values)
        )
        if len(key) == len(fields) and all(key):
            keys.add(key)
    return keys


def merge_into_workbook(
    validation: ImportValidation,
    base_path: Path,
    output_path: Path,
) -> dict[str, dict[str, int]]:
    workbook = openpyxl.load_workbook(base_path)
    summary = {sheet: {"added": 0, "skipped": 0} for sheet in IMPORT_SHEETS}
    duplicate_keys = {
        "docentes": ("docente",),
        "salas": ("campus", "codigo"),
    }

    for sheet in IMPORT_SHEETS:
        rows = validation.rows.get(sheet, [])
        if not rows:
            continue
        worksheet = _ensure_sheet(workbook, sheet)
        headers = [normalize_header(cell.value) for cell in worksheet[1]]
        existing = _existing_keys(worksheet, duplicate_keys[sheet]) if sheet in duplicate_keys else set()

        for row in rows:
            if sheet in duplicate_keys:
                key = tuple(normalize_key(row.get(field)) for field in duplicate_keys[sheet])
                if key in existing:
                    summary[sheet]["skipped"] += 1
                    continue
                existing.add(key)

            output_values = [None] * len(headers)
            for field in SHEET_COLUMNS[sheet]:
                if field in headers:
                    output_values[headers.index(field)] = _clean_value(sheet, field, row.get(field))
            worksheet.append(output_values)
            summary[sheet]["added"] += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return summary


def _write_json_atomic(data: dict[str, Any], output_path: Path) -> None:
    temporary_path = output_path.with_name(f".{output_path.name}.tmp")
    temporary_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary_path, output_path)


def _backup_file(source: Path, backup_dir: Path, timestamp: str) -> Path | None:
    if not source.exists():
        return None
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{source.stem}_{timestamp}{source.suffix}"
    shutil.copy2(source, backup_path)
    return backup_path


def main() -> int:
    root_dir = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(
        description="Valida e incorpora um novo curso; sem --apply gera apenas uma copia candidata."
    )
    parser.add_argument("input", help="Planilha de cadastro do novo curso")
    parser.add_argument(
        "--base",
        default=str(root_dir / "dados" / "planilha_base.xlsx"),
        help="Planilha-base oficial",
    )
    parser.add_argument("--output", help="Destino da copia candidata no modo de previa")
    parser.add_argument("--report", help="Destino do relatorio JSON")
    parser.add_argument("--apply", action="store_true", help="Substitui a base oficial apos backup")
    parser.add_argument("--confirm", help="Sigla do curso, obrigatoria junto com --apply")
    parser.add_argument(
        "--json-output",
        default=str(root_dir / "dados_app.json"),
        help="JSON regenerado no modo --apply",
    )
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    base_path = Path(args.base).expanduser().resolve()
    json_path = Path(args.json_output).expanduser().resolve()
    if not input_path.exists():
        parser.error(f"planilha de entrada nao encontrada: {input_path}")
    if not base_path.exists():
        parser.error(f"planilha-base nao encontrada: {base_path}")

    validation = validate_course_import(input_path, base_path)
    default_report = input_path.with_name(f"{input_path.stem}_incorporacao.json")
    report_path = Path(args.report).expanduser().resolve() if args.report else default_report
    if not validation.valid:
        write_validation_report(validation, report_path)
        print(f"Incorporacao bloqueada: {len(validation.errors)} erro(s).")
        print(f"Revise o relatorio: {report_path}")
        return 1

    if args.apply and normalize_key(args.confirm) != normalize_key(validation.course_sigla):
        parser.error(f"para aplicar, use --confirm {validation.course_sigla}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.apply:
        candidate_path = base_path.with_name(f".{base_path.stem}_{timestamp}.tmp.xlsx")
    else:
        candidate_path = (
            Path(args.output).expanduser().resolve()
            if args.output
            else root_dir
            / "dados"
            / "importacoes"
            / "saidas"
            / f"planilha_base_com_{validation.course_sigla}.xlsx"
        )

    summary = merge_into_workbook(validation, base_path, candidate_path)
    converted_data = build_json_from_excel(candidate_path)
    backups: list[str] = []

    if args.apply:
        backup_dir = base_path.parent / "backups"
        for source in (base_path, json_path):
            backup = _backup_file(source, backup_dir, timestamp)
            if backup is not None:
                backups.append(str(backup))
        os.replace(candidate_path, base_path)
        _write_json_atomic(converted_data, json_path)
        result_path = base_path
        action = "applied"
    else:
        result_path = candidate_path
        action = "preview"

    report = validation.to_report()
    report["action"] = action
    report["result_file"] = str(result_path)
    report["json_file"] = str(json_path) if args.apply else ""
    report["backups"] = backups
    report["merge"] = summary
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Curso validado: {validation.course_sigla}")
    for sheet, counts in summary.items():
        if counts["added"] or counts["skipped"]:
            print(f"- {sheet}: {counts['added']} adicionado(s), {counts['skipped']} ignorado(s)")
    if args.apply:
        print(f"Base atualizada: {base_path}")
        print(f"JSON regenerado: {json_path}")
        print(f"Backups: {base_path.parent / 'backups'}")
    else:
        print(f"Previa criada sem alterar a base: {result_path}")
        print(f"Para aplicar: python tools/merge_course_import.py \"{input_path}\" --apply --confirm {validation.course_sigla}")
    print(f"Relatorio: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
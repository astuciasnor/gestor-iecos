#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
convert_data.py
Converte a planilha_base.xlsx (modelo IECOS) em dados_app.json para o app.

- Lê abas: docentes, componentes, turmas, cursos, horarios, feriados
- Mantém nomes de colunas da sua planilha (ex.: feriados -> feriado/dia)
- Também gera aliases de compatibilidade (nome_feriado/dia_semana) para o front antigo

Requisitos:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import unicodedata
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


# -----------------------------
# Normalização / conversões
# -----------------------------
def norm_key(s: Any) -> str:
    """
    Normaliza cabeçalho de coluna para comparação:
    - lowercase
    - remove acentos
    - troca espaços por underscore
    - remove caracteres estranhos
    """
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")  # remove acento
    s = s.replace(" ", "_").replace("-", "_")
    s = "".join(ch for ch in s if ch.isalnum() or ch == "_")
    return s


def to_str(x: Any) -> str:
    return "" if x is None else str(x).strip()


def to_int(x: Any) -> int:
    if x is None or x == "":
        return 0
    try:
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return 0


def to_date_iso(x: Any) -> str:
    """Converte valores do Excel para ISO (YYYY-MM-DD)."""
    if x is None or x == "":
        return ""
    if isinstance(x, datetime):
        return x.date().isoformat()
    if isinstance(x, date):
        return x.isoformat()

    s = str(x).strip()
    if not s:
        return ""

    # já ISO
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]

    # dd/mm/yyyy
    if len(s) >= 10 and s[2] == "/" and s[5] == "/":
        dd, mm, yy = s[:2], s[3:5], s[6:10]
        return f"{yy}-{mm}-{dd}"

    return s


def is_intervalo(faixa: str) -> bool:
    """
    Detecta intervalo:
    - se contém "INTERVALO"
    - OU se a duração for <= 25 min (ex.: 10:00 - 10:20)
    """
    s = faixa.upper()
    if "INTERVALO" in s:
        return True

    try:
        parts = faixa.split("-")
        if len(parts) != 2:
            return False
        h1 = parts[0].strip()
        h2 = parts[1].strip()
        t1h, t1m = [int(p) for p in h1.split(":")]
        t2h, t2m = [int(p) for p in h2.split(":")]
        dur = (t2h * 60 + t2m) - (t1h * 60 + t1m)
        return 1 <= dur <= 25
    except Exception:
        return False


# -----------------------------
# Leitura de abas
# -----------------------------
@dataclass
class SheetSpec:
    sheet_name: str
    required: List[str]             # nomes canônicos exigidos (após aliases)
    aliases: Dict[str, str]         # header_normalizado -> canonical


def read_sheet_as_dicts(wb: openpyxl.Workbook, spec: SheetSpec) -> List[Dict[str, Any]]:
    if spec.sheet_name not in wb.sheetnames:
        raise ValueError(f"A aba '{spec.sheet_name}' não existe na planilha.")

    ws = wb[spec.sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_norm = [norm_key(h) for h in header_row]

    col_map: Dict[int, str] = {}
    for idx, h in enumerate(header_norm):
        if not h:
            continue
        canonical = spec.aliases.get(h, h)
        col_map[idx] = canonical

    present = set(col_map.values())
    missing = [c for c in spec.required if c not in present]
    if missing:
        raise ValueError(
            f"Aba '{spec.sheet_name}': faltando colunas obrigatórias: {missing}. "
            f"Encontradas: {sorted(list(present))}"
        )

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item: Dict[str, Any] = {}
        for idx, val in enumerate(row):
            key = col_map.get(idx)
            if not key:
                continue
            item[key] = val
        rows.append(item)

    return rows


def build_json_from_excel(xlsx_path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    specs = [
        SheetSpec(
            sheet_name="docentes",
            required=["docente", "unidade", "subunidade"],
            aliases={
                "docente": "docente",
                "unidade": "unidade",
                "subunidade": "subunidade",
            },
        ),
        SheetSpec(
            sheet_name="componentes",
            required=["sigla", "periodo", "codigo", "cor", "nome", "abreviacao", "ch"],
            aliases={
                "sigla": "sigla",
                "periodo": "periodo",
                "codigo": "codigo",
                "código": "codigo",
                "cor": "cor",
                "cordisciplina": "cor",      # se existir em versões antigas
                "nome": "nome",
                "abreviacao": "abreviacao",
                "ch": "ch",
            },
        ),
        SheetSpec(
            sheet_name="turmas",
            required=["sigla", "ano", "turno"],
            aliases={"sigla": "sigla", "ano": "ano", "turno": "turno"},
        ),
        SheetSpec(
            sheet_name="cursos",
            required=["sigla", "nome"],
            aliases={
                "sigla": "sigla",
                "nome": "nome",              # "Nome" normaliza para "nome"
            },
        ),
        SheetSpec(
            sheet_name="horarios",
            required=["turno", "ordem", "faixa"],
            aliases={"turno": "turno", "ordem": "ordem", "faixa": "faixa"},
        ),
        # ✅ Aqui está a correção do seu erro:
        # sua planilha tem: data, feriado, dia, tipo
        SheetSpec(
            sheet_name="feriados",
            required=["data", "feriado", "dia", "tipo"],
            aliases={
                "data": "data",
                "feriado": "feriado",
                "nome_feriado": "feriado",   # aceita também
                "dia": "dia",
                "dia_semana": "dia",         # aceita também
                "tipo": "tipo",
            },
        ),
    ]

    data: Dict[str, Any] = {
        "meta": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_file": xlsx_path.name,
        }
    }

    # Ler abas
    for spec in specs:
        data[spec.sheet_name] = read_sheet_as_dicts(wb, spec)

    # ---------- Pós-processamento (tipagem + campos derivados) ----------

    # docentes
    data["docentes"] = [
        {
            "docente": to_str(r.get("docente")),
            "unidade": to_str(r.get("unidade")),
            "subunidade": to_str(r.get("subunidade")),
        }
        for r in data["docentes"]
    ]

    # componentes
    componentes_out = []
    for r in data["componentes"]:
        sigla = to_str(r.get("sigla"))
        componentes_out.append(
            {
                # nomes iguais à planilha:
                "sigla": sigla,
                "periodo": to_str(r.get("periodo")),
                "codigo": to_str(r.get("codigo")),
                "cor": to_str(r.get("cor")),
                "nome": to_str(r.get("nome")),
                "abreviacao": to_str(r.get("abreviacao")),
                "ch": to_int(r.get("ch")),
                # compat útil no front antigo:
                "curso_sigla": sigla,
            }
        )
    data["componentes"] = componentes_out

    # alias para não quebrar código antigo
    data["disciplinas"] = componentes_out

    # cursos
    data["cursos"] = [
        {"sigla": to_str(r.get("sigla")), "nome": to_str(r.get("nome"))}
        for r in data["cursos"]
    ]

    # turmas (gera turma_id/turma_label e curso_sigla)
    turmas_out = []
    for r in data["turmas"]:
        sigla = to_str(r.get("sigla"))
        ano = to_int(r.get("ano"))
        turno = to_str(r.get("turno"))
        turma_id = f"{sigla}{ano}" if sigla and ano else ""
        turmas_out.append(
            {
                # nomes iguais à planilha:
                "sigla": sigla,
                "ano": ano,
                "turno": turno,
                # derivados úteis ao app:
                "turma_id": turma_id,
                "turma_label": turma_id,
                # compat:
                "curso_sigla": sigla,
            }
        )
    data["turmas"] = turmas_out

    # horarios (lista + dicionário por turno)
    horarios_out = []
    for r in data["horarios"]:
        turno = to_str(r.get("turno"))
        ordem = to_int(r.get("ordem"))
        faixa = to_str(r.get("faixa"))
        if not turno or not faixa:
            continue
        faixa_out = f"INTERVALO ({faixa})" if is_intervalo(faixa) else faixa
        horarios_out.append({"turno": turno, "ordem": ordem, "faixa": faixa_out})

    horarios_out.sort(key=lambda x: (x["turno"], x["ordem"]))
    data["horarios"] = horarios_out

    horarios_por_turno: Dict[str, List[str]] = {}
    for h in horarios_out:
        horarios_por_turno.setdefault(h["turno"], []).append(h["faixa"])
    data["horarios_por_turno"] = horarios_por_turno

    # feriados (reflete sua planilha: feriado/dia)
    feriados_out = []
    for r in data["feriados"]:
        nome = to_str(r.get("feriado")) or "Feriado"
        dia = to_str(r.get("dia"))
        feriados_out.append(
            {
                "data": to_date_iso(r.get("data")),
                "feriado": nome,
                "dia": dia,
                "tipo": to_str(r.get("tipo")),
                # compat opcional (para front antigo):
                "nome_feriado": nome,
                "dia_semana": dia,
            }
        )
    data["feriados"] = feriados_out

    return data


def main():
    p = argparse.ArgumentParser(description="Converte Excel (.xlsx) em JSON para o app.")
    p.add_argument("--input", required=True, help="Caminho do arquivo .xlsx (planilha_base.xlsx)")
    p.add_argument("--output", required=True, help="Caminho do arquivo .json de saída (dados_app.json)")
    args = p.parse_args()

    in_path = Path(args.input).expanduser().resolve()
    out_path = Path(args.output).expanduser().resolve()

    if not in_path.exists():
        raise SystemExit(f"Arquivo não encontrado: {in_path}")

    data = build_json_from_excel(in_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK: gerado {out_path}")
    print("Chaves:", ", ".join(data.keys()))


if __name__ == "__main__":
    main()
